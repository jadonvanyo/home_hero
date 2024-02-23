from datetime import date
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
import json
from openpyxl import Workbook
import os
from os.path import basename
import smtplib
from tabulate import tabulate


class House:
    """
    Represents a house with various attributes and methods to analyze its financial viability as an investment.
    
    Attributes:
        address (str): The street address of the house.
        sqft (float): The square footage of the house.
        price (float): The listing price of the house.
        tax (float): The yearly taxes for the house.
        rent (float): The estimated monthly rent for the house.
        property_subtype (str): Describes the type of property (e.g., single-family, duplex, etc.) and the number of units if it is a multi-family property.
        beds (str): The number of bedrooms in the house.
        baths (str): The number of bathrooms in the house, where a half bath counts as 0.5.
        description (str): A textual description of the house.
        year_built (str): The year the house was built.
        region (str): The region or area where the house is located.
        subdivision (str): The subdivision the house belongs to, if applicable.
        tax_url (str): URL to the property tax information.
        rent_url (str): URL to the rental listing or rental estimate information.
        url (str): URL to the house's listing page.
        min_rent (str): The minimum estimated rent for the house.
        max_rent (str): The maximum estimated rent for the house.
        down_payment_decimal (float): The fraction of the purchase price that must be paid upfront as a down payment.
        closing_cost_buyer_decimal (float): The fraction of the purchase price that covers the buyer's closing costs.
        closing_cost_seller_decimal (float): The fraction of the purchase price that covers the seller's closing costs.
        expected_annual_growth (float): The expected annual growth rate of the property's value.
        interest_rate (float): The annual interest rate of the mortgage.
        loan_term_yrs (int): The term of the mortgage loan in years.
        expected_repairs_monthly (float): The monthly cost of repairs as a fraction of the rent.
        expected_vacancy_monthly (float): The monthly cost associated with vacancy as a fraction of the rent.
        expected_capx_monthly (float): The monthly cost of capital expenditures as a fraction of the rent.
        expected_management_monthly (float): The monthly cost of property management as a fraction of the rent.
        insurance_rate_yearly (float): The yearly insurance rate as a fraction of the property's value.

    Methods:
        calculate_metrics(self):
            Calculates and updates various financial metrics for the house, including price per sqft, monthly insurance, down payment cost, loan amount, closing costs, monthly principle and interest payments, taxes, total operating costs, suggested total rent, and many more.
        
        email_format_html(self):
            Formats the house's data and calculated metrics into an HTML string suitable for sending as an email. This includes links to tax and rent information, a breakdown of costs, and projected financial metrics.

        email_format_plain(self):
            Formats the house's data and calculated metrics into a plain text string suitable for sending as an email. Similar to `email_format_html` but without HTML tags.
        
        featured_home_determiner(self, target_values):
            Determines whether the house meets all specified investment criteria based on the target values provided. Returns True if all criteria are met, False otherwise.

        house_excel_sheet_creator(self, wb):
            Creates a new Excel sheet in a given workbook (`wb`) and populates it with the house's data and calculated financial metrics. This method also applies formatting for better readability and analysis.

    Note:
        This class requires an external library `tabulate` for generating HTML tables and an Excel workbook object `wb` for creating Excel sheets, indicating that it should be used within a larger application context that handles Excel file manipulation and HTML content generation.
    """
    def __init__(self, config, data):
        
        self.price = float(data.get('price'))
        self.sqft = float(data.get('sqft'))
        self.tax = float(data.get('tax'))
        self.rent = float(data.get('rent'))
        self.property_subtype = data.get('property_subtype')
        self.address = data.get('address')
        self.beds = data.get('beds')
        self.baths = data.get('baths')
        self.description = data.get('description')
        self.year_built = data.get('year_built')
        self.region = data.get('region')
        self.subdivision = data.get('subdivision')
        self.tax_url = data.get('tax_url')
        self.rent_url = data.get('rent_url')
        self.url = data.get('url')
        self.min_rent = data.get('min_rent')
        self.max_rent = data.get('max_rent')
        self.down_payment_decimal = config['down_payment_decimal']
        self.closing_cost_buyer_decimal = config['closing_cost_buyer_decimal']
        self.closing_cost_seller_decimal = config['closing_cost_seller_decimal']
        self.expected_annual_growth = config['expected_annual_growth']
        self.interest_rate = config['interest_rate']
        self.loan_term_yrs = config['loan_term_yrs']
        self.expected_repairs_monthly = config['expected_repairs_monthly']
        self.expected_vacancy_monthly = config['expected_vacancy_monthly']
        self.expected_capx_monthly = config['expected_capx_monthly']
        self.expected_management_monthly = config['expected_management_monthly']
        self.insurance_rate_yearly = config['insurance_rate_yearly']
        self.calculate_metrics()


    def calculate_metrics(self):
        """
        Calculates financial and operational metrics for the house based on its attributes and configuration settings. This method updates the instance with calculated values that are crucial for assessing the property's investment viability. These calculations include but are not limited to:

        - Price per square foot: The listing price divided by the square footage of the house.
        - Monthly insurance cost: Calculated based on the yearly insurance rate and the listing price of the house.
        - Down payment cost: The upfront payment required, calculated as a percentage of the listing price.
        - Loan amount: The difference between the listing price and the down payment cost.
        - Buyer's closing costs: Additional costs required to close the deal, calculated as a percentage of the listing price.
        - Monthly principle and interest payment: The monthly payment towards the loan, calculated using the loan amount, interest rate, and loan term.
        - Monthly taxes: The property tax divided by twelve to find the monthly tax payment.
        - Total operating costs per month: Sum of principle and interest payment, monthly taxes, and monthly insurance.
        - Suggested total rent per month: Estimated rent multiplied by the number of units for multi-family properties.
        - Monthly repair expenses, capital expenditures, expected vacancy costs, and management fees: Calculated as percentages of the suggested total rent.
        - Total monthly expenses and cash flow: Sum of operating and additional monthly expenses, and the net cash flow after expenses.
        - Cash flow based on the 50% rule, total cash needed to complete the purchase, cash-on-cash return, and the 1% rule.
        - Net Operating Income (NOI), pro forma cap rate, and various annualized returns and financial projections over the loan term.

        This method leverages the property's and loan configuration attributes to perform its calculations, updating the house instance with these computed metrics for further analysis or reporting.
        """
        # Calculate the price per sqft
        self.price_per_sqft = round(self.price / self.sqft, 2)
        
        # Calculate the monthly insurance
        self.insurance_monthly = round((self.price * self.insurance_rate_yearly) / 12, 2)
        
        # Calculate the down payment needed
        self.down_payment_cost = round(self.price * self.down_payment_decimal, 2)
        
        # Calculate the loan needed
        self.loan = self.price - self.down_payment_cost
        
        # Calculate the closing costs required
        self.closing_costs = self.price * self.closing_cost_buyer_decimal
        
        # Calculate the monthly principle and interest payments
        self.principle_interest_monthly = round((self.loan * (self.interest_rate / 12) * (1 + self.interest_rate / 12) ** (self.loan_term_yrs * 12)) / ((1 + self.interest_rate / 12) ** (self.loan_term_yrs * 12) - 1), 2)
        
        # Calculate the monthly taxes
        self.taxes_monthly = round(self.tax / 12, 2)
        
        # Calculate the operating costs
        self.total_operating_costs_monthly = round(self.principle_interest_monthly + self.taxes_monthly + self.insurance_monthly, 2)
        
        # Determine how many units are contained in the property
        property_subtypes = {
            'duplex': 2,
            'triplex': 3,
            'quadplex': 4,
            'quinplex': 5
        }

        self.number_units = property_subtypes.get(self.property_subtype, 1)
            
        # Calculate the suggested total rent for the unit
        self.suggested_total_rent_monthly = round(self.number_units * self.rent, 2)
        
        # Calculate the monthly repair expenses
        self.total_repairs_monthly = round(self.suggested_total_rent_monthly * self.expected_repairs_monthly, 2)
        
        # Calculate the monthly capital expenditures
        self.total_capx_monthly = round(self.suggested_total_rent_monthly * self.expected_capx_monthly, 2)
        
        # Calculate the monthly expected vacancy
        self.total_vacancy_monthly = round(self.suggested_total_rent_monthly * self.expected_vacancy_monthly, 2)
        
        # Calculate the monthly expected management fees
        self.total_management_monthly = round(self.suggested_total_rent_monthly * self.expected_management_monthly, 2)
        
        # Calculate the total amount of monthly expanses
        self.total_expenses_monthly = self.total_operating_costs_monthly + self.total_repairs_monthly + self.total_capx_monthly + self.total_vacancy_monthly + self.total_management_monthly
        
        # Calculate the total expected monthly cash flow
        self.cash_flow_monthly = round(self.suggested_total_rent_monthly - self.total_expenses_monthly, 2)
        
        # Calculate the expected cash flow from the 50% rule
        self.cash_flow_50 = round(self.suggested_total_rent_monthly / 2 - self.principle_interest_monthly, 2)
        
        # Calculate the total cash needed to complete the deal
        self.cash_needed_total = self.down_payment_cost + self.closing_costs
        
        # Calculate the cash on cash return for the property
        self.cash_on_cash_decimal = round(self.suggested_total_rent_monthly / self.cash_needed_total, 4)
        
        # Calculate the 1% rule
        self.percent_rule_decimal = round(self.suggested_total_rent_monthly / self.price, 4)
        
        # Calculate the Net Operating Income
        self.net_operating_income = self.suggested_total_rent_monthly * 12 - (self.taxes_monthly + self.insurance_monthly + self.total_repairs_monthly + + self.total_capx_monthly + self.total_vacancy_monthly + self.total_management_monthly) * 12
        
        # Calculate the pro forma cap
        self.pro_forma_cap_decimal = round(self.net_operating_income / self.price, 4)
        
        # Declare lists for the property value, equity, loan balance, rent_growth, cashflow, profit if sold, and annualized return
        self.year = []
        self.property_value = []
        self.loan_balance = []
        self.equity = []
        self.rent_growth = []
        self.profit_if_sold = []
        self.cash_flow_yearly = []
        self.annualized_return_decimal = []
        
        # Loop though all the years of the loan term to determine yearly statistics of the property
        for x in range(self.loan_term_yrs + 1):
            self.year.append(x)
            self.property_value.append(round(self.price * (1 + self.expected_annual_growth) ** x, 2))
            self.loan_balance.append(round((self.principle_interest_monthly / (self.interest_rate / 12)) * (1 - (1 / ((1 + self.interest_rate / 12) ** (self.loan_term_yrs * 12 - x * 12)))), 2))
            self.equity.append(round(self.property_value[x] - self.loan_balance[x], 2))
            self.rent_growth.append(round(self.suggested_total_rent_monthly * (1 + self.expected_annual_growth) ** x, 2))
            self.profit_if_sold.append(round(self.property_value[x] * (1 - self.closing_cost_seller_decimal) + sum(self.cash_flow_yearly) - self.cash_needed_total - self.loan_balance[x], 2))
            self.cash_flow_yearly.append(round((self.rent_growth[x] * (1 - self.expected_repairs_monthly - self.expected_vacancy_monthly - self.expected_capx_monthly - self.expected_management_monthly) - (1 + self.expected_annual_growth) ** x * (self.insurance_monthly + self.taxes_monthly) - self.principle_interest_monthly)  * 12, 2))
            self.annualized_return_decimal.append(round((((self.profit_if_sold[x] + self.cash_needed_total) / self.cash_needed_total) ** (1 / (x + 1)) - 1), 4))
            
            
    def email_format_html(self):
        """
        Formats the house's attributes and calculated financial metrics into an HTML string, making it suitable for inclusion in an email. This method generates a visually appealing representation of the property's details and its investment analysis, including links to external resources and a table for the first five years of financial projections.

        The HTML content includes:
        - A link to the house's listing page, using the house address as the anchor text.
        - Key property details such as price, type (property subtype), layout (bedrooms, bathrooms, square footage), price per square foot, and estimated monthly rent with a link to the rent information.
        - Financial metrics including monthly operating expenses, total monthly expenses, monthly cash flow, adherence to the 1% rule, cash flow based on the 50% rule, and the estimated total cash needed for the purchase.
        - A table showing a yearly breakdown for the first five years, covering property value, loan balance, equity, expected rents, profit if sold, yearly cash flow, and annualized return percentages.
        - A brief description of the property.

        Returns:
            str: An HTML formatted string containing the house's details and financial metrics, ready to be sent as part of an email body.

        Example Usage:
            house = House(config, data)  # Assuming `config` and `data` are predefined dictionaries with property and financial info.
            email_body_html = house.email_format_html()
            # `email_body_html` can now be used as the body of an email, providing a detailed overview of the property in a rich HTML format.

        Note:
            This method is designed for generating email content that is both informative and visually engaging. The HTML string it returns should be compatible with most email clients that support HTML content.
        """
        # Establish a new list to display the annualized return as a percentage
        display_annualized_return_percent = []
        
        for year in self.annualized_return_decimal[:6]:
            year *= 100
            display_annualized_return_percent.append(year)
        
        # Add all the appropriate labels and rows for the table
        formatted_table_data = [
            ['Year'] + self.year[:6],
            ['Property Value ($)'] + self.property_value[:6],
            ['Loan Balance ($)'] + self.loan_balance[:6],
            ['Equity ($)'] + self.equity[:6],
            ['Expected Rents ($)'] + self.rent_growth[:6],
            ['Profit if Sold ($)'] + self.profit_if_sold[:6],
            ['Yearly Cash Flow ($)'] + self.cash_flow_yearly[:6],
            ['Annualized Return (%)'] + display_annualized_return_percent
        ]
        
        # Turn all the table data into HTML format
        yearly_statistics_table_html = tabulate(formatted_table_data, tablefmt='html')
        
        house_email_html = f"""
            <div>
                <h4>
                    <a href="{self.url}">{self.address}</a>
                </h4>
                <ul>
                    <li><strong>Price:</strong> ${self.price}</li>
                    <li><strong>Type:</strong> {self.property_subtype}</li>
                    <li><strong>Layout:</strong> Beds: {self.beds} Baths: {self.baths} SQFT: {self.sqft} sqft</li>
                    <li><strong>Price per SQFT:</strong> ${self.price_per_sqft}</li>
                    <li><strong>Estimated Monthly Rent:</strong> <a href="{self.rent_url}">${self.suggested_total_rent_monthly}</a></li>
                    <li><strong>Monthly Operating Expenses:</strong> ${self.total_operating_costs_monthly}</li>
                    <li><strong>Total Monthly Expenses:</strong> ${self.total_expenses_monthly}</li>
                    <li><strong>Monthly Cash Flow:</strong> ${self.cash_flow_monthly}</li>
                    <li><strong>1% Rule:</strong> {self.percent_rule_decimal * 100}%</li>
                    <li><strong>50% Rule Cash Flow:</strong> ${self.cash_flow_50}</li>
                    <li><strong>Estimated Total Cash Needed:</strong> ${self.cash_needed_total}</li>
                </ul>
                <h4>First 5 years yearly breakdown</h4>
                {yearly_statistics_table_html}
                <p>Description: {self.description}</p>
            </div>
        """
        
        return house_email_html
    
    
    def email_format_plain(self):
        """
        Formats the house's attributes and calculated financial metrics into a plain text string for email communication. This method provides a straightforward representation of the property's details and its investment analysis, suitable for environments where HTML content is not preferred or supported.

        The plain text content includes:
        - A direct URL link to the house's listing.
        - Key property details such as address, price, type (property subtype), layout (number of bedrooms and bathrooms, square footage), and the price per square foot.
        - Financial metrics including the estimated monthly rent, monthly operating expenses, total monthly expenses, monthly cash flow, the 1% rule, cash flow based on the 50% rule, and the estimated total cash needed for the purchase.
        - A summary of financial projections for the first five years, covering property value, loan balance, equity, expected rents, profit if sold, yearly cash flow, and annualized return percentages, formatted as lists.
        - A brief description of the property.

        Returns:
            str: A plain text formatted string containing the house's details and financial metrics, ready to be sent as part of an email body.

        Example Usage:
            house = House(config, data)  # Assuming `config` and `data` are predefined dictionaries with property and financial info.
            email_body_plain = house.email_format_plain()
            # `email_body_plain` can now be used as the body of an email, providing a detailed overview of the property in a simple text format.

        Note:
            This method caters to scenarios where HTML emails are not suitable or desired, offering a clear, no-frills presentation of the property's financial and physical characteristics. It's particularly useful for text-based email clients or for users who prefer or require plain text emails.
        """
        
        # Establish a new list to display the annualized return as a percentage
        display_annualized_return_percent = []
        
        for year in self.annualized_return_decimal[:6]:
            year *= 100
            display_annualized_return_percent.append(year)
        
        # Create the plain text for the house
        house_email_plain = f"""
            Link: {self.url}
            Address: {self.address}
            Price: ${self.price}
            Type: {self.property_subtype}
            Layout: Beds: {self.beds} Baths: {self.baths} SQFT: {self.sqft} sqft
            Price per SQFT: {self.price_per_sqft}
            Estimated Monthly Rent: ${self.suggested_total_rent_monthly}
            Rent URL: {self.rent_url}
            Monthly Operating Expenses: ${self.total_operating_costs_monthly}
            Total Monthly Expenses: ${self.total_expenses_monthly}
            Monthly Cash Flow: ${self.cash_flow_monthly}
            1% Rule: {self.percent_rule_decimal * 100}%
            50% Rule Cash Flow: ${self.cash_flow_50}
            Estimated Total Cash Needed: ${self.cash_needed_total}
            First 5 years yearly breakdown:
            {['Year'] + self.year[:6]}
            {['Property Value'] + self.property_value[:6]}
            {['Loan Balance'] + self.loan_balance[:6]}
            {['Equity'] + self.equity[:6]}
            {['Expected Rents'] + self.rent_growth[:6]}
            {['Profit if Sold'] + self.profit_if_sold[:6]}
            {['Yearly Cash Flow'] + self.cash_flow_yearly[:6]}
            {['Annualized Return'] + display_annualized_return_percent}
            Description: {self.description}
        """
        
        return house_email_plain
    

    def featured_home_determiner(self, target_values):
        """
        Determines whether the house meets a set of predefined investment criteria. This method compares the house's financial metrics against target values specified by the investor or analysis criteria. The target values should be provided as a dictionary where the keys correspond to specific metrics of interest, and the values represent the minimum acceptable values for those metrics.

        Parameters:
            target_values (dict): A dictionary containing the target investment criteria, with keys representing the metric names (e.g., "target_cash_flow_monthly_min", "target_percent_rule_min") and values representing the minimum acceptable values for those metrics.

        Returns:
            bool: True if the house meets or exceeds all the target criteria; False otherwise.

        Example Usage:
            target_criteria = {
                "target_cash_flow_monthly_min": 200,
                "target_percent_rule_min": 0.01,
                "target_net_operating_income_min": 2400,
                "target_pro_forma_cap_min": 0.08,
                "target_five_year_annualized_return_min": 0.1,
                "target_cash_on_cash_return_min": 0.08
            }
            house = House(config, data)  # Assuming `config` and `data` are predefined dictionaries with property and financial info.
            is_featured = house.featured_home_determiner(target_criteria)
            # `is_featured` will be True if the house meets all the target criteria, False otherwise.

        Note:
            This method is crucial for quickly identifying properties that align with an investor's specific financial goals and investment strategy. It allows for the automated screening of properties based on financial performance metrics, facilitating the investment decision-making process.
        """

        # Establish all matches between the metrics in the config file and House class   
        house_analytics_match = {
            "target_cash_flow_monthly_min": self.cash_flow_monthly,
            "target_percent_rule_min": self.percent_rule_decimal,
            "target_net_operating_income_min": self.net_operating_income,
            "target_pro_forma_cap_min": self.pro_forma_cap_decimal,
            "target_five_year_annualized_return_min": self.annualized_return_decimal[5],
            "target_cash_on_cash_return_min": self.cash_on_cash_decimal
        }
        
        # Loop through all of the potential target keys and target values available
        for key, value in house_analytics_match.items():
            # Determine if one of the keys in the potential target keys is also in the target keys
            if key in target_values:
                # Determine if the house value is equal to or greater than the target value
                if target_values[key] <= value:
                    pass
                    
                # If one key does not pass, return false
                else:
                    return False
        
        # If all keys pass, return true
        return True
    
    
    def house_excel_sheet_creator(self, wb):
        """
        Populates a new Excel sheet within a given workbook with the house's details and calculated financial metrics. This method systematically organizes key property information and investment analysis metrics into a structured Excel format, making it suitable for detailed review, comparison, and archival purposes.

        Parameters:
            wb (Workbook): An open Excel workbook object within which the new sheet will be created. This workbook should be part of a library that supports Excel file manipulation (e.g., openpyxl, xlwt).

        Returns:
            Worksheet: The newly created worksheet populated with the house's data and metrics. This includes general information about the property, financial figures, and projections that are essential for investment analysis.

        Overview of Populated Data:
            - The method creates a sheet named after the property's address, ensuring easy identification.
            - It fills the sheet with a wide array of data points, including purchase price, loan details, monthly expenses, income projections, and more, formatted appropriately for clarity and ease of understanding.
            - Financial formulas are also embedded within the sheet to dynamically calculate key investment metrics based on the populated data, enabling users to adjust certain inputs and immediately see the impact on the property's financial outlook.
            - The method applies formatting rules for better readability and analysis, such as percentage and currency formatting, where applicable.

        Example Usage:
            from openpyxl import Workbook
            wb = Workbook()
            house = House(config, data)  # Assuming `config` and `data` contain the necessary property and financial information.
            house.house_excel_sheet_creator(wb)
            wb.save('House_Analysis.xlsx')

        Note:
            This method is intended for use within a larger application that handles real estate investment analysis. It requires the caller to manage the Excel workbook's creation, saving, and other manipulations outside this method.
        """
        # Create a new name for each sheet
        sheet_name = self.address.replace(',', '').replace(' ', '-')
        
        # Create a new sheet with the specified name
        sheet = wb.create_sheet(title=sheet_name)
        
        # Format the cells for percentages and currencies
        sheet = format_excel_sheet(sheet)
        
        # Populate the sheet with the required values and formulas
        sheet['A1'] = 'Address'
        sheet['A3'] = 'Purchase Price'
        sheet['A4'] = 'Closing Costs'
        sheet['A5'] = 'Closing Costs (%)'
        sheet['A6'] = 'Annual Growth'
        sheet['A8'] = 'Loan'
        sheet['A9'] = 'Downpayment (%)'
        sheet['A10'] = 'Downpayment ($)'
        sheet['A11'] = 'Loan'
        sheet['A12'] = 'Interest Rate (%)'
        sheet['A13'] = 'Loan Term (Yrs)'
        sheet['A14'] = 'Monthly Payment'
        sheet['A16'] = 'Rental Income'
        sheet['A17'] = 'Rent'
        sheet['A19'] = 'Expenses'
        sheet['A20'] = 'Property Taxes (mo)'
        sheet['A21'] = 'Insurance (mo)'
        sheet['A22'] = 'Repairs (%-mo)'
        sheet['A23'] = 'Vacancy (%-mo)'
        sheet['A24'] = 'Capital Expenses (%-mo)'
        sheet['A25'] = 'Management Fees (%-mo)'
        sheet['A27'] = 'Year'
        sheet['A28'] = 'Property Value'
        sheet['A29'] = 'Equity'
        sheet['A30'] = 'Loan Balance'
        sheet['A31'] = 'Rent'
        sheet['A32'] = 'Cash Flow'
        sheet['A33'] = 'Profit If Sold'
        sheet['A34'] = 'Annualized Return'
        sheet['B1'] = self.address
        sheet['B3'] = self.price
        sheet['B4'] = '=B5*B3'
        sheet['B5'] = self.closing_cost_buyer_decimal
        sheet['B6'] = self.expected_annual_growth
        sheet['B9'] = self.down_payment_decimal
        sheet['B10'] = '=B3*B9'
        sheet['B11'] = '=B3-B10'
        sheet['B12'] = self.interest_rate
        sheet['B13'] = self.loan_term_yrs
        sheet['B14'] = '=(B11*(B12/12)*(1+B12/12)^(B13*12))/((1+B12/12)^(B13*12)-1)'
        sheet['B17'] = self.suggested_total_rent_monthly
        sheet['B20'] = self.taxes_monthly
        sheet['B21'] = self.insurance_monthly
        sheet['B22'] = self.expected_repairs_monthly
        sheet['B23'] = self.expected_vacancy_monthly
        sheet['B24'] = self.expected_capx_monthly
        sheet['B25'] = self.expected_management_monthly
        sheet['B27'] = 0
        sheet['B28'] = '=B3*(1+B6)^B27'
        sheet['B29'] = '=B28-B30'
        sheet['B30'] = '=B3-B10'
        sheet['B31'] = '=(B17*(1+B6)^B27)'
        sheet['B32'] = '=(B31-B31*B22-B31*B23-B31*B24-B31*B25-B21*(1+B6)^B27-B20*(1+B6)^B27-B14)*12'
        sheet['B33'] = f'=B28*(1-{self.closing_cost_seller_decimal})-E6-B30'
        sheet['B34'] = '=((B33+E6)/E6)^(1/(B27+1))-1'
        sheet['C1'] = 'Beds'
        sheet['C19'] = 'Monthly'
        sheet['C20'] = '=B20'
        sheet['C21'] = '=B21'
        sheet['C22'] = '=B22*B17'
        sheet['C23'] = '=B23*B17'
        sheet['C24'] = '=B24*B17'
        sheet['C25'] = '=B25*B17'
        sheet['C27'] = 1
        sheet['C28'] = '=B3*(1+B6)^C27'
        sheet['C29'] = '=C28-C30'
        sheet['C30'] = '=(B14/(B12/12))*(1-(1/((1+B12/12)^(B13*12-C27*12))))'
        sheet['C31'] = '=(B17*(1+B6)^C27)'
        sheet['C32'] = '=(C31-C31*B22-C31*B23-C31*B24-C31*B25-B21*(1+B6)^C27-B20*(1+B6)^C27-B14)*12'
        sheet['C33'] = f'=C28*(1-{self.closing_cost_seller_decimal})+B32-E6-C30'
        sheet['C34'] = '=((C33+E6)/E6)^(1/(C27+1))-1'
        sheet['D1'] = self.beds
        sheet['D3'] = 'Income (mo)'
        sheet['D4'] = 'Operate Cost (mo)'
        sheet['D5'] = 'Expenses (mo)'
        sheet['D6'] = 'Cash Needed'
        sheet['D8'] = 'Total CF (mo)'
        sheet['D9'] = 'CoC'
        sheet['D11'] = '1-2% Rule'
        sheet['D12'] = '50% Rule'
        sheet['D13'] = '50% Rule (CF)'
        sheet['D16'] = 'NOI (P&I not included)'
        sheet['D17'] = 'Pro Forma Cap'
        sheet['D19'] = 'Yearly'
        sheet['D20'] = '=C20*12'
        sheet['D21'] = '=C21*12'
        sheet['D22'] = '=C22*12'
        sheet['D23'] = '=C23*12'
        sheet['D24'] = '=C24*12'
        sheet['D25'] = '=C25*12'
        sheet['D27'] = 2
        sheet['D28'] = '=B3*(1+B6)^D27'
        sheet['D29'] = '=D28-D30'
        sheet['D30'] = '=(B14/(B12/12))*(1-(1/((1+B12/12)^(B13*12-D27*12))))'
        sheet['D31'] = '=(B17*(1+B6)^D27)'
        sheet['D32'] = '=(D31-D31*B22-D31*B23-D31*B24-D31*B25-B21*(1+B6)^D27-B20*(1+B6)^D27-B14)*12'
        sheet['D33'] = f'=D28*(1-{self.closing_cost_seller_decimal})+sum(B32:C32)-E6-D30'
        sheet['D34'] = '=((D33+E6)/E6)^(1/(D27+1))-1'
        sheet['E1'] = 'Baths'
        sheet['E3'] = '=B17'
        sheet['E4'] = '=B14+B20+B21'
        sheet['E5'] = '=B14+B20+B21+C22+C23+C24+C25'
        sheet['E6'] = '=B4+B10'
        sheet['E8'] = '=E3-E5'
        sheet['E9'] = '=B17/B10'
        sheet['E11'] = '=B17/B3'
        sheet['E12'] = '=B17/2'
        sheet['E13'] = '=E12-B14'
        sheet['E16'] = '=B17*12-D22-D23-D24-D25-B20*12-B21*12'
        sheet['E17'] = '=E16/B3'
        sheet['E27'] = 3
        sheet['E28'] = '=B3*(1+B6)^E27'
        sheet['E29'] = '=E28-E30'
        sheet['E30'] = '=(B14/(B12/12))*(1-(1/((1+B12/12)^(B13*12-E27*12))))'
        sheet['E31'] = '=(B17*(1+B6)^E27)'
        sheet['E32'] = '=(E31-E31*B22-E31*B23-E31*B24-E31*B25-B21*(1+B6)^E27-B20*(1+B6)^E27-B14)*12'
        sheet['E33'] = f'=E28*(1-{self.closing_cost_seller_decimal})+sum(B32:D32)-E6-E30'
        sheet['E34'] = '=((E33+E6)/E6)^(1/(E27+1))-1'
        sheet['F1'] = self.baths
        sheet['F27'] = 4
        sheet['F28'] = '=B3*(1+B6)^F27'
        sheet['F29'] = '=F28-F30'
        sheet['F30'] = '=(B14/(B12/12))*(1-(1/((1+B12/12)^(B13*12-F27*12))))'
        sheet['F31'] = '=(B17*(1+B6)^F27)'
        sheet['F32'] = '=(F31-F31*B22-F31*B23-F31*B24-F31*B25-B21*(1+B6)^F27-B20*(1+B6)^F27-B14)*12'
        sheet['F33'] = f'=F28*(1-{self.closing_cost_seller_decimal})+sum(B32:E32)-E6-F30'
        sheet['F34'] = '=((F33+E6)/E6)^(1/(F27+1))-1'
        sheet['G1'] = 'SQFT'
        sheet['G27'] = 5
        sheet['G28'] = '=B3*(1+B6)^G27'
        sheet['G29'] = '=G28-G30'
        sheet['G30'] = '=(B14/(B12/12))*(1-(1/((1+B12/12)^(B13*12-G27*12))))'
        sheet['G31'] = '=(B17*(1+B6)^G27)'
        sheet['G32'] = '=(G31-G31*B22-G31*B23-G31*B24-G31*B25-B21*(1+B6)^G27-B20*(1+B6)^G27-B14)*12'
        sheet['G33'] = f'=G28*(1-{self.closing_cost_seller_decimal})+sum(B32:F32)-E6-G30'
        sheet['G34'] = '=((G33+E6)/E6)^(1/(G27+1))-1'
        sheet['H1'] = self.sqft
        sheet['H27'] = 10
        sheet['H28'] = '=B3*(1+B6)^H27'
        sheet['H29'] = '=H28-H30'
        sheet['H30'] = '=(B14/(B12/12))*(1-(1/((1+B12/12)^(B13*12-H27*12))))'
        sheet['H31'] = '=(B17*(1+B6)^H27)'
        sheet['H32'] = '=(H31-H31*B22-H31*B23-H31*B24-H31*B25-B21*(1+B6)^H27-B20*(1+B6)^H27-B14)*12'
        sheet['I27'] = self.loan_term_yrs
        sheet['I28'] = '=B3*(1+B6)^I27'
        sheet['I29'] = '=I28-I30'
        sheet['I30'] = '=(B14/(B12/12))*(1-(1/((1+B12/12)^(B13*12-I27*12))))'
        sheet['I31'] = '=(B17*(1+B6)^I27)'
        sheet['I32'] = '=(I31-I31*B22-I31*B23-I31*B24-I31*B25-B21*(1+B6)^I27-B20*(1+B6)^I27-B14)*12'

        return sheet


def analyze_all_houses(config, data):
    """Function to analyze all the given JSON data using the House class and return a list of analyzed and error houses"""
    # Create a list with all the analyzed houses
    analyzed_houses = []
    
    # Create a list with all the houses lacking key values
    error_houses = []

    # Establish the required values to analyze a house
    required_house_values = {
        "price": lambda x: is_convertible_to_float(x) and float(x) > 0,
        "rent": lambda x: is_convertible_to_float(x) and float(x) > 0,
        "sqft": lambda x: is_convertible_to_float(x) and float(x) > 0,
        "tax": lambda x: is_convertible_to_float(x) and float(x) > 0
    }
    
    def print_house_json_error_message(key, error, json_data):
        """Function to define error messages for the homedata.json file"""
        # Error message for if a value is missing
        if error == "missing":
            print(f'"{key}" is missing for {json_data['address']} in the housedata.json file.')
        # Error message for if a value is incorrect
        elif error == "incorrect":
            print(f'"{key}" for {json_data['address']} is incorrectly entered in the housedata.json file.')
        # Error message to handle if it was not a number that was entered
        elif error == "number":
            print(f'"{key}" for {json_data['address']} is not a valid number in the housedata.json file.')
        # General error message to handle all other issues
        else:
            print("An error has occurred while verifying data from the housedata.json file.")

    # Loop through each of the houses in the dataset and create an excel sheet for that house
    for house_data in data:
        # Verify that all the data required for analyzing the house is present
        if verify_all_required_values(required_house_values, house_data, print_house_json_error_message):
            house = House(config, house_data)
            analyzed_houses.append(house)
        
        # If the calculation values for a house cannot be verified, add it to a list of error_houses
        else:
            error_houses.append(house_data['address'])
        
    return analyzed_houses, error_houses

# TODO: Rigorous testing of this function
def config_file_required_values_present(config):
    """Function to return true if all the values required for analysis are in the config file are present and accurate"""
    
    # Establish all the variables required in the config file
    required_config_values = {
        "starturls": lambda x: isinstance(x, list) and len(x) > 0,
        "down_payment_decimal": lambda x: isinstance(x, (float)) and 0 < x < 1, 
        "closing_cost_buyer_decimal": lambda x: isinstance(x, (float)) and 0 < x <= 0.25,
        "closing_cost_seller_decimal": lambda x: isinstance(x, (float)) and 0 < x <= 0.25,
        "expected_annual_growth": lambda x: isinstance(x, (float)) and 0 < x <= 2,
        "interest_rate": lambda x: isinstance(x, (float)) and 0 < x < 1,
        "loan_term_yrs": lambda x: isinstance(x, (int)) and 0 < x <= 50,
        "expected_repairs_monthly": lambda x: isinstance(x, (float)) and 0 <= x <= 0.25,
        "expected_vacancy_monthly": lambda x: isinstance(x, (float)) and 0 <= x <= 0.25,
        "expected_capx_monthly": lambda x: isinstance(x, (float)) and 0 <= x <= 0.25,
        "expected_management_monthly": lambda x: isinstance(x, (float)) and 0 <= x <= 0.25,
        "insurance_rate_yearly": lambda x: isinstance(x, (float)) and 0 <= x <= 0.25,
        "send_emails": lambda x: isinstance(x, bool),
    }
    
    def print_config_error_message(key, error, json_data=None):
        """Function to define error messages for the config file"""
        # Error message for if a value is missing
        if error == "missing":
            print(f'"{key}" is not in the config file. Please enter "{key}" in the config file.')
        # Error message for if a value is incorrect
        elif error == "incorrect":
            print(f'"{key}" is incorrectly entered in the config file. Review documentation for how to enter "{key}".')
        # General error message to handle all other issues
        else:
            print(f"An error has occurred while verifying data from the config file.")
    
    # Return False if any of the data in the config file was missing or entered incorrectly
    if not verify_all_required_values(required_config_values, config, print_config_error_message):
        return False
    
    return True
    

def create_house_analysis_excel_book(analyzed_houses, excel_filename):
    """Create an excel book given a list of analyzed House objects"""
    
    # Create a new workbook
    wb = Workbook()
    # Remove the default sheet created by openpyxl
    wb.remove(wb.active)
    
    # Loop through each of the houses in the dataset and create an excel sheet for that house
    for house in analyzed_houses:
            
        # Create the house excel sheet for the house being analyzed
        house.house_excel_sheet_creator(wb)
            
    # Save the excel file that was created
    wb.save(filename=excel_filename)

    return


def create_featured_house_email(analyzed_houses, config):
    """Function to create an email containing all of the scraped houses and some featured houses based on user request from JSON file"""
    
    # Verify that the user is looking for featured houses in their emails
    if config['featured_house_required']:
        # Generate a dictionary of target keys and values from the config file
        target_values = create_target_values_dictionary(config)
        
        # Create the beginning of the email body for all of the analyzed houses in plain text and HTML
        # email_content_plain = ""
        email_content_html = "<html>\n\t<body>\n\t\t<h2>Featured Houses:</h2>"
        
        # Loop through each of the houses in the dataset and add them to a list of analyzed houses
        for house in analyzed_houses:
            # Check to see if the analyzed house meets the investor's criteria
            if house.featured_home_determiner(target_values):
                # Add the individual house HTMl content to the total HTML content
                email_content_html += house.email_format_html()
                
                # Add the individual house plain text content to the total plain text content
                # email_content_plain += house.email_format_plain()
                
        # Close the html for the email content
        email_content_html += "\t</body>\n</html>"
        
        return email_content_html
    
    # Handle a case where the user has not asked for a featured house
    else:
        print("User did not request featured houses to be included in the email.")
        
        email_content_html = "<html>\n\t<body>\n\t\t<h3>No Featured Homes Requested</h3>\n\t</body>\n</html>"
        
        return email_content_html


def create_target_values_dictionary(config):
    """Function to return a dictionary containing all the user input target values in config"""
    
    # Dictionary to store all the target requested values
    target_values = {}
    
    # Loop through each of the target required keys and values 
    for key, value in config.items():
        # Determine if a value exits in config for a target key
        if value is not None:
            target_values[f'{key}'] = value
            
    return target_values  


def delete_file(file_path):
    """Function to delete a given file"""
    # Try to delete the file
    try:
        os.remove(file_path)
        print(f"File {file_path} has been successfully deleted.")
        
    # Handle issues if the file is not found
    except FileNotFoundError:
        print(f"File not found: {file_path}")
        
    # Handle issues if permission is needed to delete the file
    except PermissionError as pe:
        print(f"Permission error deleting file {file_path}: {pe}")
        
    # Handle any additional exceptions
    except Exception as e:
        print(f"Error deleting file {file_path}: {e}")


def email_config_file_required_values_present(config):
    """Function to verify and return a dictionary of all the required email values and return false otherwise"""
    
    # Load the email config file for verification
    email_config = load_json(config['email_config_file_path'])
    
    # Return false if the email config file fails to load
    if not email_config:
        return False
    
    # Establish all the variables required for email in the config file
    required_config_email_values = {
        "email_config_file_path": lambda x: isinstance(x, str),
        "send_error_emails": lambda x: isinstance(x, bool),
        "featured_house_required": lambda x: isinstance(x, bool)
    }
    
    def print_email_config_error_message(key, error, json_data=None):
        """Function to define error messages for the email config file"""
        # Error message for if a value is missing
        if error == "missing":
            print(f'"{key}" is not in the config file. Please enter "{key}" in the config file.')
        # Error message for if a value is incorrect
        elif error == "incorrect":
            print(f'"{key}" is incorrectly entered in the config file. Review documentation for how to enter "{key}".')
        # General error message to handle all other issues
        else:
            print(f"An error has occurred while verifying data from the config file.")
    
    # Verify that all the required values for email in the config file exist and are valid
    if not verify_all_required_values(required_config_email_values, config, print_email_config_error_message):
        return False

    # Verify that all the required information in the email config file is present
    elif not verify_email_config_file(config):
        return False
    
    # Determine whether the user wants to include featured houses in their email
    elif config['featured_house_required']:
        if not verify_config_file_target_values(config):
            return False
        
    # Return all the required information to send an email if all checks pass
    return email_config


def format_excel_sheet(sheet):
    """Format an excel sheet for the house data"""
    
    # Apply the percentage format to all percentage cells
    percentage_format = "#0.00%"
    sheet['B5'].number_format = percentage_format
    sheet['B6'].number_format = percentage_format
    sheet['B9'].number_format = percentage_format
    sheet['B12'].number_format = percentage_format
    sheet['B22'].number_format = percentage_format
    sheet['B23'].number_format = percentage_format
    sheet['B24'].number_format = percentage_format
    sheet['B25'].number_format = percentage_format
    sheet['B34'].number_format = percentage_format
    sheet['C34'].number_format = percentage_format
    sheet['D34'].number_format = percentage_format
    sheet['E9'].number_format = percentage_format
    sheet['E11'].number_format = percentage_format
    sheet['E17'].number_format = percentage_format
    sheet['E34'].number_format = percentage_format
    sheet['F34'].number_format = percentage_format
    sheet['G34'].number_format = percentage_format
    
    # Apply currency format to all currency cells
    currency_format = '"$"#,###,##0.00'
    sheet['B3'].number_format = currency_format
    sheet['B4'].number_format = currency_format
    sheet['B10'].number_format = currency_format
    sheet['B11'].number_format = currency_format
    sheet['B14'].number_format = currency_format
    sheet['B17'].number_format = currency_format
    sheet['B20'].number_format = currency_format
    sheet['B21'].number_format = currency_format
    sheet['B28'].number_format = currency_format
    sheet['B29'].number_format = currency_format
    sheet['B30'].number_format = currency_format
    sheet['B31'].number_format = currency_format
    sheet['B32'].number_format = currency_format
    sheet['B33'].number_format = currency_format
    sheet['C22'].number_format = currency_format
    sheet['C23'].number_format = currency_format
    sheet['C24'].number_format = currency_format
    sheet['C25'].number_format = currency_format
    sheet['C28'].number_format = currency_format
    sheet['C29'].number_format = currency_format
    sheet['C30'].number_format = currency_format
    sheet['C31'].number_format = currency_format
    sheet['C32'].number_format = currency_format
    sheet['C33'].number_format = currency_format
    sheet['D20'].number_format = currency_format
    sheet['D21'].number_format = currency_format
    sheet['D22'].number_format = currency_format
    sheet['D23'].number_format = currency_format
    sheet['D24'].number_format = currency_format
    sheet['D25'].number_format = currency_format
    sheet['D28'].number_format = currency_format
    sheet['D29'].number_format = currency_format
    sheet['D30'].number_format = currency_format
    sheet['D31'].number_format = currency_format
    sheet['D32'].number_format = currency_format
    sheet['D33'].number_format = currency_format
    sheet['E12'].number_format = currency_format
    sheet['E16'].number_format = currency_format
    sheet['E28'].number_format = currency_format
    sheet['E29'].number_format = currency_format
    sheet['E30'].number_format = currency_format
    sheet['E31'].number_format = currency_format
    sheet['E32'].number_format = currency_format
    sheet['E33'].number_format = currency_format
    sheet['F28'].number_format = currency_format
    sheet['F29'].number_format = currency_format
    sheet['F30'].number_format = currency_format
    sheet['F31'].number_format = currency_format
    sheet['F32'].number_format = currency_format
    sheet['F33'].number_format = currency_format
    sheet['G28'].number_format = currency_format
    sheet['G29'].number_format = currency_format
    sheet['G30'].number_format = currency_format
    sheet['G31'].number_format = currency_format
    sheet['G32'].number_format = currency_format
    sheet['G33'].number_format = currency_format
    sheet['H28'].number_format = currency_format
    sheet['H29'].number_format = currency_format
    sheet['H30'].number_format = currency_format
    sheet['H31'].number_format = currency_format
    sheet['H32'].number_format = currency_format
    sheet['H33'].number_format = currency_format
    sheet['I28'].number_format = currency_format
    sheet['I29'].number_format = currency_format
    sheet['I30'].number_format = currency_format
    sheet['I31'].number_format = currency_format
    sheet['I32'].number_format = currency_format
    sheet['I33'].number_format = currency_format
    
    return sheet


def is_convertible_to_float(s):
    try:
        float(s)
        return True
    except ValueError:
        return False


def load_json(json_path):
    """Load a configuration file with sensitive or variable information"""
    # Try to open the json file
    try:
        with open(json_path, 'r') as json_file:
            json_data = json.load(json_file)
        return json_data
    # Handle errors if the email config file is not found
    except:
        print(f"An error occurred while trying to load '{json_path}'. Verify that the target json file name matches, that the file exists, and is complete.")
        return


def send_error_email(error_message, config, required_email_values):
    """Function to send a custom error message to a user if any issues occur that they cannot see and exit the program"""
    
    # Verify that the user wants emails
    if config['send_emails']:
        
        # Verify that the user wants error messages emailed
        if config['send_error_emails']:
            
            # Setup the MIME
            message = MIMEMultipart()
            message['From'] = required_email_values['sender_address']
            message['To'] = required_email_values['receiver_address']
            message['Subject'] = 'An error has occurred'   # The subject line

            # Attach the plain text to also be sent with the email
            message.attach(MIMEText(error_message, 'plain'))
            
            # Create SMTP session for sending the mail
            session = smtplib.SMTP('smtp.gmail.com', 587) 
            session.starttls() # enable security
            session.login(required_email_values['sender_address'], required_email_values['password']) # login with mail_id and password
            session.sendmail(required_email_values['sender_address'], required_email_values['receiver_address'], message.as_string()) # Send an email with the excel file attached
            session.quit()
            print('Mail Sent')
        
    return


def send_featured_house_email(excel_filename, email_content_html, config, required_email_values):
    """Function to send an email containing the spreadsheet and any featured houses to a specified user"""
        
    # Verify that the user wants emails
    if config['send_emails']:
        # Setup the MIME
        message = MIMEMultipart()
        message['From'] = required_email_values['sender_address']
        message['To'] = required_email_values['receiver_address']
        message['Subject'] = f'Houses analyzed - {str(date.today())}'   # The subject line

        # Attach the HTML to also be sent with the email
        message.attach(MIMEText(email_content_html, 'html'))

        # Try to open the excel file to send in an email
        try:
            # Open the excel file and include it as an attachment for the email
            with open(excel_filename, 'rb') as file:
                part = MIMEApplication(file.read(), Name=basename(excel_filename))
                part["Content-Disposition"] = f'attachment; filename="{basename(excel_filename)}"'
                message.attach(part)
        except FileNotFoundError:
            print("It appears the excel file was not created. Verify the excel file is being created before sending the email.")
            return

        # Create SMTP session for sending the mail
        session = smtplib.SMTP('smtp.gmail.com', 587) 
        session.starttls() # enable security
        session.login(required_email_values['sender_address'], required_email_values['password']) # login with mail_id and password
        session.sendmail(required_email_values['sender_address'], required_email_values['receiver_address'], message.as_string()) # Send an email with the excel file attached
        session.quit()
        print('Mail Sent')
    
    return


def verify_all_required_values(required_values, json_data, error_messages=None):
    """Function to return True if all the required values are included and within specified ranges for any JSON data"""
    # Variable to track if the data in the json_data file has been entered incorrectly
    file_correct = True
    
    # Loop through each of the keys and values required for the json_data file
    for key, value in required_values.items():
        # Verify that each of the keys exists in the json_data file
        if key in json_data:
            
            # Handles if the value at a key in the json_data file is null
            if json_data[key] is None:
                error_messages(key, 'missing', json_data)
                file_correct = False
                
            # Handles if the value in the json_data file does not fall in the limits of the required values
            elif not value(json_data[key]):
                error_messages(key, 'incorrect', json_data)
                file_correct = False
        
        # Handles if a key is missing in the json_data file
        else:
            error_messages(key, 'missing', json_data)
            file_correct = False
    
    # Return file_correct for verification if all values were present
    return file_correct

        
def verify_config_file_target_values(config):
    """Function to test and return a boolean expression representing if the config file contains valid target values"""
    
    # Establish all the potential target variables required
    required_target_values = {
        "target_cash_flow_monthly_min": lambda x: isinstance(x, (int, float)),
        "target_percent_rule_min": lambda x: isinstance(x, (float)) and 0 < x < 1,
        "target_net_operating_income_min": lambda x: isinstance(x, (int, float)),
        "target_pro_forma_cap_min": lambda x: isinstance(x, (float)) and 0 < x < 1,
        "target_five_year_annualized_return_min": lambda x: isinstance(x, (float)) and 0 < x < 1,
        "target_cash_on_cash_return_min": lambda x: isinstance(x, (float)) and 0 < x < 1
    }
    
    # Establish a list to keep track of if any target values were established
    target_values_established = []
    
    # Loop through each of the target required keys and values 
    for key, value in required_target_values.items():
        # Determine if a value exits in config for a target key
        if config[key] is not None:
            # If a target value does not fit the required criteria return an error message and append false
            if not value(config[key]):
                print(f'"{key}" was entered incorrectly in the config file. Refer to the documentation on how to enter "{key}".')
                target_values_established.append(False)
            
            # If the value does pass the test, append true
            else:
                target_values_established.append(True)
    
    # Return an error if the user selected to have featured houses, but they did not set any parameters
    if not target_values_established:
        print("'featured_house_required' was selected, but no target values were established. Refer to the documentation on how to use 'featured_house_required'.")
        return False
    
    # Return false if any of the selected target values failed verification
    elif False in target_values_established:
        return False
    
    # If all variables passed, return true
    else:
        return True
        
        
def verify_email_config_file(config):
    """Function to return a boolean expression to verify that the email config file works and that it contains all the required values to send emails"""

    # Establish all the variables required for the email config file
    required_email_config_values = {
        "sender_address": lambda x: isinstance(x, str) and "@" in x,
        "receiver_address": lambda x: isinstance(x, str) and "@" in x,
        "password": lambda x: isinstance(x, str)
    }
    
    def print_email_config_error_message(key, error, json_data=None):
        """Function to define error messages for the config file"""
        # Error message for if a value is missing
        if error == "missing":
            print(f'"{key}" is not in the email config file. Please enter "{key}" in the email config file.')
        # Error message for if a value is incorrect
        elif error == "incorrect":
            print(f'"{key}" is incorrectly entered in the email config file. Review documentation for how to enter "{key}".')
        # General error message to handle all other issues
        else:
            print(f"An error has occurred while verifying data from the config file.")
    
    # Pull all the email data from a separate config file
    email_config = load_json(config['email_config_file_path'])

    # Verify email config file can be opened
    if not email_config:
        return False
    
    # TODO: Update the email files to run from config only
    # TODO: Eliminate all other checks for the the email or config data
    if not verify_all_required_values(required_email_config_values, email_config, print_email_config_error_message):
        return False
         
    # Return true if all checks have passed without returning false
    return True